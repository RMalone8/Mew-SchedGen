import customtkinter as ctk
from PIL import Image
import openpyxl as pyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import random
import shutil
from pathlib import Path
import numpy as np
from datetime import datetime
from template_create import template_generator
import os

# Sets the appearance mode of the application
# "System" sets the appearance same as that of the system
ctk.set_appearance_mode("dark-blue")        
 
# Sets the color of the widgets 
# Supported themes: green, dark-blue, blue
ctk.set_default_color_theme("dark-blue")
 
appWidth=1125
appHeight=650

save_file_name = "past_activity_data.xlsx"
activity_names = []
activity_categs = []
previous_schedules = []

# Loading our saved data... if we have any
if os.path.exists(save_file_name):
    data_book = pyxl.open(save_file_name)
    data_sheet = data_book.active
    for i in range(data_sheet["A1"].value):
        activity_names.append(data_sheet[f"A{i+2}"].value.split("&/&")[0])
        activity_categs.append(data_sheet[f"A{i+2}"].value.split("&/&")[1])
    for i in range(data_sheet["B1"].value):
        letter = get_column_letter(i+2)
        previous_sched = []
        for j in range(data_sheet[f"{letter}2"].value):
            previous_sched.append(data_sheet[f"{letter}{j+3}"].value.split("&/&"))
        previous_schedules.append(previous_sched)
else:
    data_book = Workbook()
    data_sheet = data_book.active
    data_sheet["A1"].value = 0
    data_sheet["B1"].value = 0
    data_book.save(save_file_name)

extra_act = [f"Enter Activity" for i in range(30 - len(activity_names))]
extra_categs = ["NA" for _ in range(30 - len(activity_names))]
activity_names += extra_act
activity_categs += extra_categs

class MyFrame(ctk.CTkScrollableFrame):
    def __init__(self, master, actnames: list, **kwargs):
        super().__init__(master, **kwargs)
        #Initialize Checkboxes
        self.checkboxVars = [ctk.StringVar(value="off") for _ in range(len(actnames))]
        self.checkboxspan = [ctk.StringVar(value="off") for _ in range(len(actnames))]
        self.checkboxboy = [ctk.StringVar(value="on") for _ in range(len(actnames))]
        self.checkboxgirl = [ctk.StringVar(value="on") for _ in range(len(actnames))]
        self.checkboxsimul= [ctk.StringVar(value="off") for _ in range(len(actnames))]
        self.choices = [ctk.CTkCheckBox(self, text=" ", variable=self.checkboxVars[i], onvalue="on", offvalue="off") for i in range(len(actnames))]
        for i in range(len(self.choices)):
            self.choices[i].grid(row=i, column=3, padx=0, pady=20, sticky="ew")

        self.activity_entries = []
        self.activity_buttons = []

        self.entry = [ctk.CTkEntry(self, placeholder_text=actnames[i], placeholder_text_color=("white", "white")) for i in range(len(actnames))]
        for i in range(len(self.entry)):
            self.entry[i].grid(row=i, column=4, columnspan=1, padx=5, pady=10, sticky="ew")

        for i in range(len(self.checkboxspan)):
            self.checkboxspan[i] = ctk.CTkCheckBox(self, text="Span Two Periods", variable=self.checkboxspan[i], onvalue="on", offvalue="off")
            self.checkboxspan[i].grid(row=i, column=5, padx=20, pady=20, sticky="ew")
            if "Double" in activity_categs[i]:
                self.checkboxspan[i].select()
        for i in range(len(self.checkboxsimul)):
            self.checkboxsimul[i] = ctk.CTkCheckBox(self, text="Simultaneous", variable=self.checkboxsimul[i], onvalue="on", offvalue="off")
            self.checkboxsimul[i].grid(row=i, column=6, padx=20, pady=20, sticky="ew")
            if "Simul" in activity_categs[i]:
                self.checkboxsimul[i].select()
        for i in range(len(self.checkboxgirl)):
            self.checkboxgirl[i] = ctk.CTkCheckBox(self, text="Girl's Side", variable=self.checkboxgirl[i], onvalue="on", offvalue="off")
            self.checkboxgirl[i].grid(row=i, column=7, padx=20, pady=20, sticky="ew")
        for i in range(len(self.checkboxboy)):
            self.checkboxboy[i] = ctk.CTkCheckBox(self, text="Boy's Side", variable=self.checkboxboy[i], onvalue="on", offvalue="off")
            self.checkboxboy[i].grid(row=i, column=8, padx=20, pady=20, sticky="ew")
            if "JustBoy" in activity_categs[i]:
                self.checkboxgirl[i].deselect()
            elif "JustGirl" in activity_categs[i]:
                self.checkboxboy[i].deselect()

class Memory(ctk.CTkScrollableFrame):
    def __init__(self, master: list, **kwargs):
        super().__init__(master, **kwargs)

        def save_previous_schedule():
            temp=1
            self.save_schedule_button=ctk.CTkButton(self,text="Save Schedule", corner_radius=32, fg_color="#0000FF",hover_color="#33BFFF",command=save_previous_schedule)
            self.save_schedule_button.grid(row=0,column=0,columnspan=1,padx=10,pady=10,sticky="ew")


# Create App class
class App(ctk.CTk):
# Layout of the GUI will be written in the init itself
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.title("Mew: Schedule Generator")    

        self.geometry(f"{appWidth}x{appHeight}")    
 
        self.num_of_acts_used = 12 # Initial number of actvities!
        self.just_generated_schedule = []

        self.my_frame = MyFrame(master=self, actnames=activity_names, width=850, height=300)
        self.my_frame.grid(row=2, column=1,columnspan=12, padx=20, pady=20)

        self.error_message = "Click to Generate a Schedule"

        # Schedule Name/Title
        self.nameLabel = ctk.CTkLabel(self,
                                      text="Schedule Name")
        self.nameLabel.grid(row=0, column=0,
                            padx=3, pady=10,
                            sticky="ew")
        self.nameEntry = ctk.CTkEntry(self,
                         placeholder_text="Enter Schedule Name")
        self.nameEntry.grid(row=0, column=1,
                            columnspan=3, padx=3,
                            pady=10, sticky="ew")


        #Alter Number of Activities
        def alternumacts():
            self.num_of_acts_used=int(self.numberactsentry.get())
            if(int(self.numberactsentry.get())>30):
                self.error_message = "Cannot Select More Than 30 Activities"
                self.messageDisplayButton.configure(text=self.error_message, fg_color="#FF6961", hover_color="#FF6961")
            else: 
                for i in range(self.num_of_acts_used):
                    self.my_frame.checkboxVars[i].set("on")
                for i in range(self.num_of_acts_used, 30):
                    self.my_frame.checkboxVars[i].set("off")
                self.error_message = " "
                self.messageDisplayButton.configure(text=self.error_message, fg_color="#FFFFFF", hover_color="#FFFFFF")

        self.numberacts = ctk.CTkLabel(self,
                                       text="Number of Activities")
        self.numberacts.grid(row=0, column=5,
                             padx=5, pady=10,
                             sticky="ew")
        self.numberactsentry = ctk.CTkEntry(self,
                         placeholder_text="Enter Here")
        self.numberactsentry.grid(row=0, column=6,
                            columnspan=1, padx=5,
                            pady=10, sticky="ew")
        self.numberactsbutton=ctk.CTkButton(self,text="Select", corner_radius=32, fg_color="#0000FF",hover_color="#33BFFF",command=alternumacts)
        self.numberactsbutton.grid(row=0,column=7,columnspan=1,padx=15,pady=10,sticky="ew")        
         
        def testrun():
            self.messageDisplayButton.configure(text=self.error_message)
        
        def deselectall():
            for i in range(len(self.my_frame.choices)):
                self.my_frame.checkboxVars[i].set("off")
            
        def saveSchedule():
            # Logging the Schedules into the Sheet
            if self.just_generated_schedule:
                letter = get_column_letter(data_sheet["B1"].value+2)
                data_sheet[f"{letter}2"].value = len(self.just_generated_schedule)
                for i in range(len(self.just_generated_schedule)):
                    data_sheet[f"{letter}{i+3}"].value = "&/&".join(self.just_generated_schedule[i])
                data_sheet["B1"].value += 1
                data_book.save(save_file_name)
                previous_schedules.append(self.just_generated_schedule)
                self.error_message = "Schedule Saved!"
                self.messageDisplayButton.configure(text=self.error_message, fg_color="#77DD77", hover_color="#77DD77")
            else:
                self.error_message = "No Previous Schedules to Save"
                self.messageDisplayButton.configure(text=self.error_message, fg_color="#FF6961", hover_color="#FF6961")
        def eraseSchedules():
            for i in range(data_sheet["B1"].value):
                letter = get_column_letter(i+2)
                for j in range(data_sheet[f"{letter}2"].value):
                    data_sheet[f"{letter}{j+3}"].value = None
                data_sheet[f"{letter}2"].value = None
            data_sheet["B1"].value = 0
            data_book.save(save_file_name)
            self.error_message = "Successully Deleted Previous Schedule Memory"
            self.messageDisplayButton.configure(text=self.error_message, fg_color="#77DD77", hover_color="#77DD77")

        def run(): #button press
            total_acts = len(self.my_frame.choices)
            opts = [True if self.my_frame.checkboxVars[i].get() == "on" else False for i in range(total_acts)]
            actnames=[self.my_frame.entry[i].get() if self.my_frame.entry[i].get() else self.my_frame.entry[i].cget("placeholder_text") for i in range(total_acts)]
            self.num_of_acts_used = sum(opts)
            act_dict = []
            #Sort out the categories
            categs = ["All" if self.my_frame.checkboxboy[i].get() == self.my_frame.checkboxgirl[i].get() else "JustBoy" if self.my_frame.checkboxboy[i].get() == "on" else "JustGirl" for i in range(total_acts)]
            categs = [categs[i]+"Simul" if self.my_frame.checkboxsimul[i].get() == "on" else categs[i] for i in range(total_acts)]
            categs = [categs[i]+"Double" if self.my_frame.checkboxspan[i].get() == "on" else categs[i] for i in range(total_acts)]
            time_period = "morning" if self.checkboxmorning.get() == "on" else "afternoon" if self.checkboxafternoon.get() == "on" else "wholeday"

            # Emptying the data sheet for new previous schedule
            length = data_sheet[f"A1"].value
            for i in range(length):
                data_sheet[f"A{i+2}"].value = None
            data_sheet["A1"].value = self.num_of_acts_used
            # Creating our act_dict for the generator and storing those activities!
            data_row = 2
            for i in range(total_acts):
                    if opts[i]:
                        act_dict.append({"name": actnames[i], "type": categs[i]})
                        data_sheet[f"A{data_row}"].value = actnames[i] + "&/&" + categs[i]
                        data_row += 1
            data_book.save(save_file_name)

            global specs

            if time_period == "morning":
                specs = {"rows": ["4", "5", "6", "7", "8", "10", "11", "12", "13", "14"],
                        "cols": ['C', 'D', 'E'],
                        "blocked_coords": []} 
            elif time_period == "afternoon":
                specs = {"rows": ["4", "5", "6", "7", "8", "10", "11", "12", "13", "14"],
                        "cols": ['D', 'F', 'G', 'H'],
                        "blocked_coords": [[2,0], [2,1], [2,2], [2,5], [2,6], [2,7]]} 
            else:
                specs = {"rows": ["4", "5", "6", "7", "8", "10", "11", "12", "13", "14"],
                        "cols": ['C', 'D', 'E', 'G', 'I', 'J', 'K'],
                        "blocked_coords": [[5,0], [5,1], [5,2], [5,5], [5,6], [5,7]]} 

            def getSchedule(b_groups: int, g_groups: int, activities: list, sheet_specs: dict):
                count = 0
                result = generateSchedule(b_groups= b_groups, g_groups= g_groups, activities= activities, sheet_specs=sheet_specs)
                print(result)
                while (not result and count < 200) or compareAgainstPrevious(proposed_sched=result, previous_scheds=previous_schedules) > 5:
                    result = generateSchedule(b_groups= b_groups, g_groups= g_groups, activities= activities, sheet_specs=sheet_specs)
                    count +=1
                #if not result and count >= 100:
                #    self.error_message = "No Sir, Too Similar!!"
                #    self.messageDisplayButton.configure(text=self.error_message, fg_color="#FF6961", hover_color="#FF6961")
                if count < 200 and result:
                    # Sending Produced Schedule to Downloads
                    output_path=Path.home()/'Downloads'/f'{self.nameEntry.get().replace(" ", "_")}.xlsx'#f'Generated_Schedule_{datetime.now().month}_{datetime.now().day}_{datetime.now().year}.xlsx'
                    shutil.copyfile('new_sheet.xlsx',output_path)
                    print(f"Here is how similar it is to previous schedules (210 meaning identical): {compareAgainstPrevious(proposed_sched=result, previous_scheds=previous_schedules)}")
                if result:
                    self.error_message = "Schedule Generated!"
                    self.messageDisplayButton.configure(text=self.error_message, fg_color="#77DD77", hover_color="#77DD77")
                else:
                    self.error_message = "Failure to Generate Schedule"
                    self.messageDisplayButton.configure(text=self.error_message, fg_color="#FF6961", hover_color="#FF6961")
                
                return result

            def generateSchedule(b_groups: int, g_groups: int, activities: list, sheet_specs: dict):
                periods = len(sheet_specs["cols"])
                groups = b_groups + g_groups
                new_activity = ""
                global override_coords
                override_coords = []

                book = template_generator(time_period)
                sheet = book.active

                # Checking if a valid schedule can even be made!
                if groups != len(sheet_specs["rows"]):
                    return False
                if groups > len(activities) or periods > len(activities):
                    return False
                
                global schedule
                schedule = [["" for group in range(groups)] for period in range(periods)]

                for i in range(groups):
                    for j in range(periods):
                        while True:
                            if not schedule[j][i]: # If there isn't already something there
                                new_activity = chooseValidActivity(group=i, period=j, groups=groups, periods=periods, activities=activities)
                                if new_activity:
                                    schedule[j][i] = new_activity["name"]
                                    sheet[sheet_specs["cols"][j]+sheet_specs["rows"][i]].value = new_activity["name"]
                                    break
                                else:
                                    return False
                            else:
                                for coords in override_coords:
                                    if [j, i] == coords["coords"]: # Are these coordinates to be skipped?
                                        sheet[sheet_specs["cols"][j]+sheet_specs["rows"][i]].value = schedule[j][i]
                                        override_coords.remove(coords)
                                        break 
                                break

                # Sending the finished schedule...
                sheet["A1"].value = self.nameEntry.get()
                book.save('new_sheet.xlsx') 
                return schedule

            def chooseValidActivity(group: int, period: int, groups: int, periods: int, activities: list) -> str:
                avail_activities = activities.copy()
                for act in avail_activities[:]:
                    if act["name"] in schedule[period]:
                        avail_activities.remove(act)
                    elif act["name"] in np.transpose(schedule)[group]:
                        avail_activities.remove(act)
                    elif "Girl" in act["type"] and group < groups/2:
                        avail_activities.remove(act)
                    elif "Boy" in act["type"] and group >= groups/2:
                        avail_activities.remove(act)

                    #Simulateous Conditions    
                    elif "Simul" in act["type"] and (group == groups - 1 or group == groups/2 - 1):
                        avail_activities.remove(act)
                    elif "Simul" in act["type"] and [period, group+1] in specs['blocked_coords']:
                        avail_activities.remove(act)
                    elif "Simul" in act["type"] and schedule[period][group+1]:
                        avail_activities.remove(act)

                    #Double Conditions
                    elif "Double" in act["type"] and (period == periods - 1 or act["name"] in schedule[period+1]):
                        avail_activities.remove(act)
                    elif "Double" in act["type"] and abs(ord(specs['cols'][period]) - ord(specs['cols'][period+1])) > 1:
                        avail_activities.remove(act)
                    elif "Double" in act["type"] and [period+1, group] in specs['blocked_coords']:
                        avail_activities.remove(act)
                    elif "Double" in act["type"] and schedule[period+1][group]:
                        avail_activities.remove(act)

                if [period, group] in specs['blocked_coords']:
                    new_act = {"name": "", "type": "None"}
                elif avail_activities:
                    new_act = random.choice(avail_activities)
                else:
                    return False
                if "Simul" in new_act["type"] and "Double" in new_act["type"]:
                    schedule[period][group+1] = new_act["name"]
                    override_coords.append({"coords": [period, group+1]})
                    schedule[period+1][group] = new_act["name"]
                    override_coords.append({"coords": [period+1, group]})
                    schedule[period+1][group+1] = new_act["name"]
                    override_coords.append({"coords": [period+1, group+1]})
                elif "Simul" in new_act["type"]:
                    schedule[period][group+1] = new_act["name"]
                    override_coords.append({"coords": [period, group+1]})
                elif "Double" in new_act["type"]:
                    schedule[period+1][group] = new_act["name"]
                    override_coords.append({"coords": [period+1, group]})
                return new_act
            
            def compareAgainstPrevious(proposed_sched: list, previous_scheds: list):
                if not proposed_sched:
                    return False

                same_acts = 0
                same_times = 0
                total_score = 0
                proposed = np.array(proposed_sched).transpose()

                for previous in previous_scheds:
                    if len(proposed_sched) == len(previous):
                        for new_row, old_row,  in zip(proposed, np.array(previous).transpose()):
                            new_row = [act.replace(" ", "") for act in new_row] # Stripped all spaces in case of random mistakes space-wise
                            old_row = [act.replace(" ", "") for act in old_row]
                            same_times = sum([3 for new_act, old_act in zip(new_row, old_row) if new_act == old_act])
                            same_acts = sum([1 for old_act in old_row if old_act in new_row])
                        total_score = same_times + same_acts if same_acts + same_times > total_score else total_score
                return total_score

            self.just_generated_schedule = getSchedule(b_groups=5, g_groups=5, activities=act_dict, sheet_specs=specs)
            print(self.just_generated_schedule)
            
        # Functions to maintain one enabled timeframe checkbox
        def reset_checkboxes_morn():
            if self.checkboxmorning.get() == "on":
                self.checkboxafternoon.set("off")
                self.checkboxwholeday.set("off")
        def reset_checkboxes_afternoon():
            if self.checkboxafternoon.get() == "on":
                self.checkboxmorning.set("off")
                self.checkboxwholeday.set("off")
        def reset_checkboxes_wholeday():
            if self.checkboxwholeday.get() == "on":
                self.checkboxafternoon.set("off")
                self.checkboxmorning.set("off")

        # Creating the checkboxes
        self.checkboxmorning=ctk.StringVar(value="off")
        self.checkboxafternoon=ctk.StringVar(value="off")
        self.checkboxwholeday=ctk.StringVar(value="on")
        self.choicemorning = ctk.CTkCheckBox(self,
                            text="Morning",
                            variable=self.checkboxmorning,
                            onvalue="on",
                            offvalue="off",
                            command=reset_checkboxes_morn)                               
        self.choicemorning.grid(row=7, column=4,
                          padx=5, pady=5,
                          sticky="ew")
        
        self.choiceafternoon = ctk.CTkCheckBox(self,
                            text="Afternoon",
                            variable=self.checkboxafternoon,
                            onvalue="on",
                            offvalue="off",
                            command=reset_checkboxes_afternoon)                             
        self.choiceafternoon.grid(row=7, column=5,
                          padx=5, pady=5,
                          sticky="ew")

        self.choicewholeday = ctk.CTkCheckBox(self,
                            text="Whole Day",
                            variable=self.checkboxwholeday,
                            onvalue="on",
                            offvalue="off",
                            command=reset_checkboxes_wholeday)                               
        self.choicewholeday.grid(row=7, column=6,
                          padx=5, pady=5,
                          sticky="ew")

        #Make Spreadsheet Button
        self.deselectButton=ctk.CTkButton(self,text="Deselect All Activities", corner_radius=32, fg_color="#0000FF",hover_color="#33BFFF",command=deselectall)
        self.deselectButton.grid(row=1,column=4,columnspan=3,padx=10,pady=10,sticky="ew")

        self.saveButton=ctk.CTkButton(self,text="Save Previous", text_color="#000000", corner_radius=32, fg_color="#77DD77",hover_color="#B9FEB9",command=saveSchedule)
        self.saveButton.grid(row=1,column=3,columnspan=1,padx=5,pady=5,sticky="ew")

        self.eraseButton=ctk.CTkButton(self,text="Erase Memory", text_color="#000000", corner_radius=32, fg_color="#FF6961",hover_color="#f69697",command=eraseSchedules)
        self.eraseButton.grid(row=1,column=7,columnspan=1,padx=5,pady=5,sticky="ew")

        self.durationPromptButton=ctk.CTkButton(self,text="Select Duration of Schedule", text_color="#000000", corner_radius=32, fg_color="#FFFFFF",hover_color="#FFFFFF",command=testrun)
        self.durationPromptButton.grid(row=6,column=4,columnspan=3,padx=10,pady=10,sticky="ew")

        self.generateResultsButton=ctk.CTkButton(self,text="Generate Schedule", corner_radius=32, fg_color="#0000FF",hover_color="#33BFFF",command=run)
        self.generateResultsButton.grid(row=8,column=4,columnspan=3,padx=10,pady=10,sticky="ew")

        self.messageDisplayButton=ctk.CTkButton(self,text=" ", text_color="#000000", corner_radius=32, fg_color="#FFFFFF",hover_color="#FFFFFF",command=testrun)
        self.messageDisplayButton.grid(row=9,column=4,columnspan=3,padx=10,pady=10,sticky="ew")
      
        #my_image=ctk.CTkImage(Image.open('images/CampIHCLogo2.png'),size=(150,150))
        #my_image_label=ctk.CTkLabel(self, image=my_image, width=200, height=200)
        #my_image_label.grid(row=7,column=9,columnspan=1,padx=10,pady=10,sticky="ew")
        #my_image_label.grid_propagate(False)

if __name__ == "__main__":
    app = App()
    # Runs the app
    app.mainloop()