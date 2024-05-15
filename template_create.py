from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import openpyxl.styles as xlsty

def template_generator(type: str):
    book = Workbook()
    sheet = book.active

    # Fonts and Alignment
    font_title = xlsty.Font(name="Arial", size=20, bold=True)
    font_b = xlsty.Font(name="Arial", size=10, bold=True)
    font = xlsty.Font(name="Arial", size=10)
    align = xlsty.Alignment(horizontal="center", vertical="center")

    # Colors
    col_red = xlsty.PatternFill(patternType="solid", fgColor="00FFA0A0")
    col_blue = xlsty.PatternFill(patternType="solid", fgColor="00A0A0FF")
    col_light_blue = xlsty.PatternFill(patternType="solid", fgColor="00ADD8E6")
    col_pink = xlsty.PatternFill(patternType="solid", fgColor="00FF69B4")
    col_light_pink = xlsty.PatternFill(patternType="solid", fgColor="00FDD8EF")
    col_black = xlsty.PatternFill(patternType="solid", fgColor="00000000")
    col_grey = xlsty.PatternFill(patternType="solid", fgColor="00505050")

    # Borders
    black_border_side = xlsty.Side(border_style="thin", color="00000000")
    black_border = xlsty.Border(left=black_border_side, right=black_border_side,
                                top=black_border_side, bottom=black_border_side)
    period_titles = ["Period 1\n9:50-10:40am", "Period 2\n10:50-11:30am", "Period 3\n11:40am-12:30pm", 
                     "12:30-2:10pm", "Period 4\n2:10-3:10pm", "3:10-3:40pm", 
                     "Period 5\n3:40-4:40pm", "Period 6\n4:50-5:50pm", "Evening Activity"]
    period_shift = 0
    merges = []
    sheet_length = 0
    column_lunch = ""
    column_milk = ""



    if type == "wholeday":
        merges = ["A1:K2", "A4:A8", "A10:A14", "F4:F8", "F10:F14", "H4:H8", "H10:H14", "A9:K9", "A15:K15"]
        sheet_length = 12
        column_lunch = "F"
        column_milk = "H"
        period_shift = 0
    elif type == "morning":
        merges = ["A1:F2", "A4:A8", "A10:A14", "F4:F8", "F10:F14", "A9:F9", "A15:F15"]
        sheet_length = 7
        column_lunch = "F"
        column_milk = ""
        period_shift = 0
    else:
        merges = ["A1:H2", "A4:A8", "A10:A14", "C4:C8", "C10:C14", "E4:E8", "E10:E14", "A9:H9", "A15:H15"]
        sheet_length = 9
        column_lunch = "C"
        column_milk = "E"
        period_shift = 3

    # Merge necessary cells
    for merge in merges:
        sheet.merge_cells(merge)
        sheet[merge][0][0].font = font_b
        sheet[merge][0][0].alignment = align

    # Iterating through each excel row to style
    for i in range(3,15):
        sheet["B"+ str(i)].font = font_b
        sheet["B"+ str(i)].alignment = align

    # So inefficient but here we go:
    sheet["A1"].fill = col_red
    sheet["A1"].font = font_title

    sheet["A4"].value = "Boy's Side"
    sheet["A4"].fill = col_blue
    sheet["B3"].value = "Division"
    sheet["B4"].value = "Rookies"
    sheet["B5"].value = "Explorers"
    sheet["B6"].value = "Rangers"
    sheet["B7"].value = "Trailblazers"
    sheet["B8"].value = "Graduates"

    sheet["A10"].value = "Girl's Side"
    sheet["A10"].fill = col_pink
    sheet["B10"].value = "Rookies"
    sheet["B11"].value = "Explorers"
    sheet["B12"].value = "Rangers"
    sheet["B13"].value = "Trailblazers"
    sheet["B14"].value = "Graduates"

    sheet["A3"].fill = col_red
    sheet["A3"].border = black_border

    sheet["B3"].fill = col_red
    sheet["B3"].border = black_border
    sheet["B3"].value = "Division"
    sheet["B3"].font = font_b

    for i in range(sheet_length-3):
        sheet[get_column_letter(i+3)+"3"].fill = col_red
        sheet[get_column_letter(i+3)+"3"].border = black_border
        sheet[get_column_letter(i+3)+"3"].font = font_b
        sheet[get_column_letter(i+3)+"3"].alignment = align
        sheet[get_column_letter(i+3)+"3"].value = period_titles[i+period_shift]

    # Place lunch/m&c at proper times! (if applicable)
    sheet[column_lunch+"4"].value = "Lunch/\nRest Hour"
    sheet[column_lunch+"4"].font = font
    sheet[column_lunch+"10"].value = "Lunch/\nRest Hour"
    sheet[column_lunch+"10"].font = font

    if column_milk:
        sheet[column_milk+"4"].value = "Milk & Coookies\n@ Pavillion"
        sheet[column_milk+"4"].font = font
        sheet[column_milk+"10"].value = "Milk & Coookies\n@ Pavillion"
        sheet[column_milk+"10"].font = font

    for i in range(1,sheet_length):
        sheet.column_dimensions[get_column_letter(i)].width = 20
    sheet.row_dimensions[1].height = 20
    sheet.row_dimensions[3].height = 40

    # Coloring in the proper cells!
    for row in range(4,9):
        for col in range(2,sheet_length):
            sheet[get_column_letter(col)+str(row)].fill = col_light_blue
            sheet[get_column_letter(col)+str(row)].border = black_border
            sheet[get_column_letter(col)+str(row)].alignment = align

    for row in range(10,15):
        for col in range(2,sheet_length):
            sheet[get_column_letter(col)+str(row)].fill = col_light_pink
            sheet[get_column_letter(col)+str(row)].border = black_border
            sheet[get_column_letter(col)+str(row)].alignment = align


    # Don't forget to black out the the necessary period 6 blocks!
    if type == "wholeday":
        sheet["J4"].fill = col_black
        sheet["J5"].fill = col_black
        sheet["J6"].fill = col_black
        sheet["J10"].fill = col_black
        sheet["J11"].fill = col_black
        sheet["J12"].fill = col_black
    elif type == "afternoon":
        sheet["G4"].fill = col_black
        sheet["G5"].fill = col_black
        sheet["G6"].fill = col_black
        sheet["G10"].fill = col_black
        sheet["G11"].fill = col_black
        sheet["G12"].fill = col_black

    # And the other grey borders
    sheet["A9"].fill = col_grey
    sheet["A15"].fill = col_grey

    return book



